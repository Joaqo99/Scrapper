from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
#me quedé haciendo la función para crear la lista de guitarras. La idea es que el programa evalúe por sí solo cuales celdas están en blanco y q él
#mismo cree el rango, si necesidad de que yo le diga hasta donde es. Luego de eso, el programa va a crear una variable para el rango horizontal
#y otra para el vertical. La función va a crear listas dentro de listas, ejemplo: [[Jackson, Kelly, rhoads], [ibanez, s], [Gibson, Les paul, explorer, sg, firebird]]
#el primer elemento de cada sublista debe ser la marca, y el resto los modelos

def obtenerLista(file):
    wb = load_workbook(file)
    ws = wb["Buscar"]
    tabla = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == "None":
            break
        fila = list(row)
        while str(fila[-1]) == "None":
            fila.pop()
        tabla.append(fila)
        
    return tabla

def escribir(file, encabezados):
    wb = load_workbook(file)
    ws = wb["Resultados"]
    
    ws.append(encabezados)

    
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('guitarras.xlsx')
ws = wb.active

#print(ws.rows)
#Así se lee automaticamente:
#for row in ws.rows:
#    for cell in row:
#        print(cell.value)

for row in ws.iter_rows(min_row=2, values_only=True):
    if str(row[0]) == "None":
        break
    fila = list(row)
    while str(fila[-1]) == "None":
        fila.pop()
    print(fila)

